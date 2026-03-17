# ==========================================================
#  Proyecto: Herramienta para el análisis multitemporal
#            de cuerpos de agua
#  Autor: Xavier Escobar Arteus18
#  ==========================================================
import sys
import os
from PIL import Image 

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QGraphicsDropShadowEffect, QSizeGrip,
    QMessageBox, QFileDialog, QVBoxLayout,QSlider,  QLabel
)
from PySide6.QtCore import QFile, QStandardPaths
from PySide6.QtGui import QColor
from PySide6 import QtCore, QtWidgets
from PySide6.QtUiTools import QUiLoader

import pandas as pd
import rasterio
import numpy as np

import matplotlib
matplotlib.use("Qt5Agg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

from rasterio.warp import calculate_default_transform, reproject, Resampling
from skimage import morphology
import openpyxl


class CargarImagen(QMainWindow):
    def __init__(self, parent=None, codigo=None):
        super().__init__(None)

        # ---------------- Estado compartido de la clase ----------------
        self.codigo = str(codigo).strip() if codigo is not None else ""
        self.ruta_imagen = ""
        self.satelite = ""          # "Sentinel" o "Landsat"
        self.water_mask = None      # máscara final agua (bool)
        self.cloud_mask = None      # máscara nubes (bool)
        self.ndwi = None            # ndwi (float)
   
        # ---------------- Ventana ----------------
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        self.setWindowOpacity(1)
        self.clickPosition = None
        self.parent_window = parent
        if parent:
            self.setGeometry(parent.geometry())
        
        
          # UI
        self.cargar_ui()
        self.configurar_botones()
        self.agregar_sombra_widgets()
        self.configurar_redimensionamiento()
        self.configurar_eventos_movimiento()
        self.configurar_slider_Mnubes()
        self.configurar_slider_Ndwi()
        
        self.ventana.btn_Aplicar.setEnabled(False)
        self.ventana.btn_Guardar.setEnabled(False)
        
        # ---------------- db_path fijo ----------------
        docs = QStandardPaths.writableLocation(QStandardPaths.DocumentsLocation)
        self.db_path = os.path.join(docs, "CuerposDeAgua", "BD", "BaseCuerposDeAgua.xlsx")
        print(f"Ruta del archivo Excel: {self.db_path}")

    # =================================================================
    # UI
    # =================================================================
    def cargar_ui(self):
        ui_file = QFile(r"C:\Users\LENOVO\Documents\CuerposDeAgua\UI\MenuSubirImagen.ui")
        ui_file.open(QFile.ReadOnly)
        loader = QUiLoader()
        self.ventana = loader.load(ui_file)
        ui_file.close()

        if self.ventana is None:
            print("No se pudo cargar la interfaz.")
            sys.exit(-1)

        self.setCentralWidget(self.ventana)
        
    def configurar_slider_Mnubes(self):
        
        self.lbl_Nubes = self.ventana.findChild(QLabel, "lbl_Nubes")
        
        
        self.ventana.sld_Nubes.setMinimum(0)  # Mínimo valor del slider
        self.ventana.sld_Nubes.setMaximum(100)   # Máximo valor del slider
        self.ventana.sld_Nubes.setTickInterval(5)  # Intervalo de ticks
        self.ventana.sld_Nubes.setValue(0)
        self.ventana.sld_Nubes.setTickPosition(QSlider.TicksBelow)  # Posición de los ticks
        self.ventana.sld_Nubes.valueChanged.connect(self.actualizar_valor_Mnubes)
        
    
    def actualizar_valor_Mnubes(self):
        # Obtener el valor del slider en el rango de -1 a 1
        self.valor_slider_nube = self.ventana.sld_Nubes.value() / 100.0  # Convertir de -100 a 100 a -1 a 1
        # Actualizar el texto de la etiqueta lbl_Ndwi con el nuevo valor
        self.lbl_Nubes.setText(f" {self.valor_slider_nube:.2f}")

    def configurar_slider_Ndwi(self):
        
        self.lbl_Ndwi = self.ventana.findChild(QLabel, "lbl_Ndwi")
        
        
        self.ventana.sld_Ndwi.setMinimum(-100)  # Mínimo valor del slider
        self.ventana.sld_Ndwi.setMaximum(100)   # Máximo valor del slider
        self.ventana.sld_Ndwi.setTickInterval(5)  # Intervalo de ticks
        self.ventana.sld_Ndwi.setValue(-100)
        self.ventana.sld_Ndwi.setTickPosition(QSlider.TicksBelow)  # Posición de los ticks
        self.ventana.sld_Ndwi.valueChanged.connect(self.actualizar_valor_Ndwi)
        
    
    def actualizar_valor_Ndwi(self):
        # Obtener el valor del slider en el rango de -1 a 1
        self.valor_slider = self.ventana.sld_Ndwi.value() / 100.0  # Convertir de -100 a 100 a -1 a 1
        # Actualizar el texto de la etiqueta lbl_Ndwi con el nuevo valor
        self.lbl_Ndwi.setText(f" {self.valor_slider:.2f}")
    
    def configurar_botones(self):
        self.ventana.btn_Normal.hide()

        self.ventana.btn_Minimizar.clicked.connect(self.control_bt_minimizar)
        self.ventana.btn_Maximizar.clicked.connect(self.control_bt_maximizar)
        self.ventana.btn_Normal.clicked.connect(self.control_bt_normal)
        self.ventana.btn_Cerrar.clicked.connect(self.control_bt_Cerrar)

        self.ventana.btn_Subir.clicked.connect(self.cargar_imagen)
        self.ventana.btn_Guardar.clicked.connect(self.guardar_mascara_y_resultados)
        self.ventana.btn_Aplicar.clicked.connect(self.btnAplicar)
        self.ventana.btn_Cancelar.clicked.connect(self.CerrarImagen)
        

        self.ventana.Cbox_Satelite.currentIndexChanged.connect(self.cargar_bandas)

    def agregar_sombra_widgets(self):
        widgets = [
            self.ventana.btn_Subir, self.ventana.btn_Aplicar,
            self.ventana.btn_Guardar, self.ventana.btn_Cancelar,
            self.ventana.Cbox_Satelite
        ]
        for widget in widgets:
            self.sombra_frame(widget)

    def configurar_redimensionamiento(self):
        self.gripSize = 10
        self.grip = QSizeGrip(self)
        self.grip.resize(self.gripSize, self.gripSize)
        self.grip.move(self.ventana.width() - self.gripSize, self.ventana.height() - self.gripSize)

    def configurar_eventos_movimiento(self):
        self.ventana.frame_superior.mousePressEvent = self.mousePressEvent
        self.ventana.frame_superior.mouseMoveEvent = self.mouseMoveEvent

    def control_bt_minimizar(self):
        self.showMinimized()

    def control_bt_normal(self):
        self.showNormal()
        self.ventana.btn_Normal.hide()
        self.ventana.btn_Maximizar.show()

    def control_bt_maximizar(self):
        self.showMaximized()
        self.ventana.btn_Maximizar.hide()
        self.ventana.btn_Normal.show()

    def control_bt_Cerrar(self):
        self.close()
        QApplication.quit()

    def mousePressEvent(self, event):
        if event.buttons() == QtCore.Qt.LeftButton:
            self.clickPosition = event.globalPos()
            event.accept()

    def mouseMoveEvent(self, event):
        if not self.isMaximized() and event.buttons() == QtCore.Qt.LeftButton:
            if self.clickPosition is not None:
                click_pos = self.clickPosition if isinstance(self.clickPosition, QtCore.QPoint) else self.clickPosition.toPoint()
                global_pos = event.globalPos()
                delta = global_pos - click_pos
                self.move(self.pos() + delta)
                self.clickPosition = global_pos
            event.accept()

        if event.globalPos().y() <= 10:
            self.showMaximized()
            self.ventana.btn_Maximizar.hide()
            self.ventana.btn_Normal.show()
        else:
            self.showNormal()
            self.ventana.btn_Normal.hide()
            self.ventana.btn_Maximizar.show()

    def sombra_frame(self, frame):
        sombra = QGraphicsDropShadowEffect(self)
        sombra.setBlurRadius(40)
        sombra.setXOffset(8)
        sombra.setYOffset(8)
        sombra.setColor(QColor(39, 173, 194, 255))
        frame.setGraphicsEffect(sombra)

    def resizeEvent(self, event):
        rect = self.rect()
        self.grip.move(rect.right() - self.gripSize, rect.bottom() - self.gripSize)

    # =================================================================
    # Selección y carga de imagen / bandas
    # =================================================================
    def cargar_imagen(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar Imagen", "",
            "Imágenes (*.tif *.tiff *.jpg *.png);;Todos los archivos (*)",
            options=options
        )
        if file_path:
            self.ruta_imagen = file_path
            print(f"Imagen seleccionada: {self.ruta_imagen}")
        
        nombre_archivo = os.path.basename(self.ruta_imagen)
        lbl_NomL = self.ventana.findChild(QtWidgets.QLabel, "lbl_Fecha")
      

        if lbl_NomL:
            lbl_NomL.setText(f" {nombre_archivo}")
        
        self.ventana.btn_Aplicar.setEnabled(True)
        self.ventana.btn_Guardar.setEnabled(True)
        self.ventana.sld_Ndwi.setValue(-30)  
        self.ventana.sld_Nubes.setValue(10)
        self.cargar_bandas()

    def cargar_bandas(self):
        if not self.ruta_imagen:
            return

        self.satelite = self.ventana.Cbox_Satelite.currentText().strip()
        print(f"Satélite seleccionado: {self.satelite}")

        if self.satelite == "Sentinel":
            self.cargar_bandas_sentinel()
        elif self.satelite == "Landsat":
            self.cargar_bandas_landsat()

    def cargar_bandas_sentinel(self):
        with rasterio.open(self.ruta_imagen) as src:
            blue = src.read(1).astype("float32")
            green = src.read(2).astype("float32")
            red = src.read(3).astype("float32")
            nir = src.read(4).astype("float32")
            swir1 = src.read(5).astype("float32")

        rgb = np.dstack([red, green, blue])
        p2, p98 = np.percentile(rgb, (2, 98))
        self.rgb_norm = np.clip((rgb - p2) / (p98 - p2), 0, 1)
        self.mostrar_imagen_procesada(self.rgb_norm)

        # Guardar cloud_mask como atributo
        thr_swir = self.valor_slider_nube
        self.cloud_mask, self.ndwi = self._calcular_ndwi_y_mascara_nubes(green, nir, swir1, thr_swir)
        self._dibujar_en_widget(self.ventana.wdg_Ndwi, self.ndwi, cmap="RdYlGn", vmin=-1, vmax=1)
        self._dibujar_en_widget(self.ventana.wdg_Mnubes, swir1, cmap="magma")

        # Guardar water_mask como atributo
        umbral = self.valor_slider 
        self.water_mask = self._procesar_cuerpo_agua(self.ndwi, umbral)
        self.mostrar_mascara_agua(self.water_mask)
        

    def cargar_bandas_landsat(self):
        with rasterio.open(self.ruta_imagen) as src:
            # según tu idea: blue=2, green=3, red=4, nir=5
            blue = src.read(1).astype("float32")
            green = src.read(2).astype("float32")
            red = src.read(3).astype("float32")
            nir = src.read(4).astype("float32")
            swir1 = src.read(5).astype("float32") 
        rgb = np.dstack([red, green, blue])
        p2, p98 = np.percentile(rgb, (2, 98))
        self.rgb_norm = np.clip((rgb - p2) / (p98 - p2), 0, 1)
        self.mostrar_imagen_procesada(self.rgb_norm)
        
        thr_swir = self.valor_slider_nube
        self.cloud_mask, self.ndwi = self._calcular_ndwi_y_mascara_nubes(green, nir, swir1, thr_swir)

        self._dibujar_en_widget(self.ventana.wdg_Ndwi, self.ndwi, cmap="RdYlGn", vmin=-1, vmax=1)
        self._dibujar_en_widget(self.ventana.wdg_Mnubes, swir1, cmap="magma")
        umbral= self.valor_slider 
        self.water_mask = self._procesar_cuerpo_agua(self.ndwi, umbral)
        self.mostrar_mascara_agua(self.water_mask)
        

    # =================================================================
    # Visualización en widgets
    # =================================================================
    def mostrar_imagen_procesada(self, imagen):
        
        # 1) Asegura layout único
        if self.ventana.wdg_Original.layout() is None:
            self.ventana.wdg_Original.setLayout(QVBoxLayout())
        layout = self.ventana.wdg_Original.layout()

        # 2) Limpia lo anterior (canvas viejo)
        self._clear_layout(self.ventana.wdg_Original)
        
        fig, ax = plt.subplots(figsize=(6, 6))
        ax.imshow(imagen)
        ax.axis("off")
        fig.subplots_adjust(left=0, right=1, top=1, bottom=0)

        canvas = FigureCanvas(fig)
        layout.addWidget(canvas)
        canvas.draw()
        plt.close(fig)
    
    
    def _clear_layout(self, widget):
        layout = widget.layout()
        if layout is None:
            return
        while layout.count():
            item = layout.takeAt(0)
            w = item.widget()
            if w is not None:
                w.deleteLater()

    def _clear_and_get_layout(self, widget: QtWidgets.QWidget) -> QVBoxLayout:
        layout = widget.layout()
        if layout is None:
            layout = QVBoxLayout(widget)
            layout.setContentsMargins(0, 0, 0, 0)
            layout.setSpacing(0)
            widget.setLayout(layout)

        while layout.count():
            item = layout.takeAt(0)
            w = item.widget()
            if w is not None:
                w.deleteLater()

        return layout

    def _dibujar_en_widget(self, widget: QtWidgets.QWidget, img, cmap=None, vmin=None, vmax=None):
        layout = self._clear_and_get_layout(widget)

        fig = Figure()
        ax = fig.add_axes([0.02 , 0.10, 0.80, 0.80])
        ax.set_axis_off()

        if cmap is None:
            im = ax.imshow(img)
        else:
            cm = plt.get_cmap(cmap).copy()
            cm.set_bad(alpha=0)  # NaN -> transparente
            im = ax.imshow(img, cmap=cm, vmin=vmin, vmax=vmax)

        cbar = fig.colorbar(im, ax=ax, orientation="vertical", fraction=0.03, pad=0.04)
        cbar.ax.tick_params(labelsize=5)
        fig.subplots_adjust(left=0.05, right=0.85, top=1, bottom=0)

        canvas = FigureCanvas(fig)
        canvas.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        layout.addWidget(canvas)
        canvas.draw()
        plt.close(fig)

    def mostrar_mascara_agua(self, water_mask):
        
        # 1) Asegura layout único
        if self.ventana.wdg_Magua.layout() is None:
            self.ventana.wdg_Magua.setLayout(QVBoxLayout())
        layout = self.ventana.wdg_Magua.layout()

        # 2) Limpia lo anterior (canvas viejo)
        self._clear_layout(self.ventana.wdg_Magua)
        
        fig, ax = plt.subplots(figsize=(6, 6))
        ax.imshow(water_mask, cmap="Blues")
        ax.axis("off")
        fig.subplots_adjust(left=0, right=1, top=1, bottom=0)

        canvas = FigureCanvas(fig)
        layout.addWidget(canvas)
        canvas.draw()
        plt.close(fig)

    # =================================================================
    # Cálculos (NDWI, máscara agua)
    # =================================================================
    def _calcular_ndwi_y_mascara_nubes(self, green, nir, swir1, thr_swir):
        eps = 1e-6
        cloud_mask = swir1 > thr_swir
        ndwi = (green - nir) / (green + nir + eps)
        ndwi = ndwi.astype("float32")
        ndwi[cloud_mask] = np.nan
        return cloud_mask, ndwi

    def _procesar_cuerpo_agua(self, ndwi, umbral):
        print(umbral)
        water_mask = ndwi > umbral
        water_clean = morphology.remove_small_objects(water_mask, min_size=300)
        water_clean = morphology.remove_small_holes(water_clean, area_threshold=200)
        water_clean = morphology.binary_closing(water_clean, footprint=morphology.disk(2))
        return water_clean

    # =================================================================
    # Reproyección (Método 1) para área en m²
    # =================================================================
    def _get_crs_transform_y_grid_utm(self, src, dst_crs="EPSG:32717"):
        transform, width, height = calculate_default_transform(
            src.crs, dst_crs, src.width, src.height, *src.bounds
        )
        return dst_crs, transform, width, height

    def _reproyectar_banda(self, src, arr_src, dst_crs, dst_transform, dst_width, dst_height, resampling):
        dst = np.empty((dst_height, dst_width), dtype=arr_src.dtype)
        reproject(
            source=arr_src,
            destination=dst,
            src_transform=src.transform,
            src_crs=src.crs,
            dst_transform=dst_transform,
            dst_crs=dst_crs,
            resampling=resampling
        )
        return dst

    # =================================================================
    # Excel I/O
    # =================================================================
    def _leer_hoja(self, ruta: str, hoja: str, columnas: list[str]) -> pd.DataFrame:
        if os.path.exists(ruta):
            try:
                df = pd.read_excel(ruta, sheet_name=hoja)
                for c in columnas:
                    if c not in df.columns:
                        df[c] = ""
                return df[columnas]
            except Exception:
                pass
        return pd.DataFrame(columns=columnas)

    # =================================================================
    # Guardar máscara + NDTI + métricas (con reproyección UTM si hace falta)
    # =================================================================
    def guardar_mascara_y_resultados(self):
        # -------- Validaciones --------
        if not self.codigo:
            QMessageBox.warning(self, "Advertencia", "No se recibió 'codigo'.")
            return
        if not self.ruta_imagen:
            QMessageBox.warning(self, "Advertencia", "Primero selecciona una imagen.")
            return
        if self.water_mask is None:
            QMessageBox.warning(self, "Advertencia", "Primero calcula la máscara (elige satélite y carga bandas).")
            return

        codigo = str(self.codigo).strip()
        nombre_archivo = os.path.basename(self.ruta_imagen)
        fecha_archivo = os.path.splitext(nombre_archivo)[0]  # fecha = nombre del archivo

        # -------- Leer MEDICIONES y validar duplicado --------
        cols_med = ["codigo", "fecha", "area", "ndti_media", "ndti_median", "p10", "p25", "p75", "p90", "ruta_mask", "ruta_ndti"]
        df_med = self._leer_hoja(self.db_path, "MEDICIONES", cols_med)
        if not df_med.empty:
            df_med["codigo"] = df_med["codigo"].astype(str).str.strip()
            df_med["fecha"] = df_med["fecha"].astype(str).str.strip()

        if not df_med[(df_med["codigo"] == codigo) & (df_med["fecha"] == fecha_archivo)].empty:
            QMessageBox.warning(self, "Advertencia", f"Ya existe un registro con código '{codigo}' y fecha '{fecha_archivo}'.")
            return

        # -------- Leer LAGUNAS y obtener ruta_proyecto --------
        df_lag = self._leer_hoja(self.db_path, "LAGUNAS", ["codigo", "ruta_proyecto"])
        if df_lag.empty:
            QMessageBox.warning(self, "Advertencia", "La hoja LAGUNAS está vacía o no existe.")
            return

        df_lag["codigo"] = df_lag["codigo"].astype(str).str.strip()
        fila = df_lag[df_lag["codigo"] == codigo]
        if fila.empty:
            QMessageBox.warning(self, "Advertencia", f"No se encontró el código '{codigo}' en LAGUNAS.")
            return

        ruta_proyecto = str(fila["ruta_proyecto"].values[0]).strip()
        if not ruta_proyecto or ruta_proyecto.lower() == "nan":
            QMessageBox.warning(self, "Advertencia", "ruta_proyecto está vacía para ese código.")
            return

        # -------- Crear carpeta de salida --------
        carpeta_mascara = os.path.join(ruta_proyecto, "Procesamiento_Imágnes")
        os.makedirs(carpeta_mascara, exist_ok=True)
        
        carpeta_wmask = os.path.join(carpeta_mascara, "MascaraDeAgua")
        os.makedirs(carpeta_wmask, exist_ok=True)
        
        carpeta_ndti = os.path.join(carpeta_mascara, "Mascarandti")
        os.makedirs(carpeta_ndti, exist_ok=True)
        
        carpeta_rgb = os.path.join(carpeta_mascara, "ImgOriginal")
        os.makedirs(carpeta_rgb, exist_ok=True)

        out_mask_path = os.path.join(carpeta_wmask, f"{fecha_archivo}.tif")
        out_ndti_path = os.path.join(carpeta_ndti, f"{fecha_archivo}.tif")
        
        # Convertir rgb_norm de [0,1] a [0,255] y guardar la imagen
        rgb_norm_uint8 = (self.rgb_norm * 255).astype(np.uint8)

        # Crear una imagen PIL desde el array numpy
        image_pil = Image.fromarray(rgb_norm_uint8)

        # Guardar la imagen como PNG (puedes cambiar la extensión a .tiff, .jpg, etc.)
        output_path = os.path.join(carpeta_rgb, f"{fecha_archivo}.tif")
        image_pil.save(output_path)

        # -------- Procesamiento + reproyección UTM (método 1) si hace falta --------
        try:
            with rasterio.open(self.ruta_imagen) as src:
                print("CRS original:", src.crs)

                sat = (self.satelite or "").strip()

                if src.crs is not None and src.crs.is_projected:
                    # Raster ya proyectado
                    dst_crs = src.crs
                    dst_transform = src.transform
                    dst_width, dst_height = src.width, src.height

                    # Bandas locales (NO self) para NDTI
                    if sat == "Landsat":
                        green = src.read(2).astype("float32")
                        red = src.read(3).astype("float32")
                    else:
                        green = src.read(2).astype("float32")
                        red = src.read(3).astype("float32")

                    water_mask_dst = self.water_mask.astype(np.uint8)
                    cloud_mask_dst = self.cloud_mask.astype(np.uint8) if self.cloud_mask is not None else None

                else:
                    # Raster en lat/lon -> reproyectar a UTM zona Ecuador
                    dst_crs, dst_transform, dst_width, dst_height = self._get_crs_transform_y_grid_utm(src, dst_crs="EPSG:32717")
                    print("⚠️ Raster en lat/lon → reproyectando a UTM", dst_crs)

                    if sat == "Landsat":
                        green_src = src.read(2).astype("float32")
                        red_src = src.read(3).astype("float32")
                    else:
                        green_src = src.read(2).astype("float32")
                        red_src = src.read(3).astype("float32")

                    # reproyectar bandas
                    green = self._reproyectar_banda(src, green_src, dst_crs, dst_transform, dst_width, dst_height, Resampling.nearest)
                    red = self._reproyectar_banda(src, red_src, dst_crs, dst_transform, dst_width, dst_height, Resampling.nearest)

                    # reproyectar máscaras
                    water_mask_dst = self._reproyectar_banda(
                        src, self.water_mask.astype(np.uint8), dst_crs, dst_transform, dst_width, dst_height, Resampling.nearest
                    ).astype(np.uint8)

                    cloud_mask_dst = None
                    if self.cloud_mask is not None:
                        cloud_mask_dst = self._reproyectar_banda(
                            src, self.cloud_mask.astype(np.uint8), dst_crs, dst_transform, dst_width, dst_height, Resampling.nearest
                        ).astype(np.uint8)

                # -------- Área en m² (ahora correcto porque es CRS proyectado) --------
                pixel_area = abs(dst_transform.a * dst_transform.e)  # e suele ser negativo
                print(float(pixel_area))
                area_agua = float(np.sum(water_mask_dst > 0)) * float(pixel_area)
                area_agua=area_agua/ 1000000

                # -------- NDTI en grid destino --------
                eps = 1e-6
                ndti = (red - green) / (red + green + eps)

                if cloud_mask_dst is not None:
                    ndti[cloud_mask_dst > 0] = np.nan

                ndti_water = np.where(water_mask_dst > 0, ndti, np.nan).astype("float32")

                # -------- Estadísticas SOLO AGUA --------
                vals = ndti_water[np.isfinite(ndti_water)]
                if vals.size == 0:
                    ndti_mean = ndti_median = np.nan
                    p10 = p25 = p75 = p90 = np.nan
                else:
                    ndti_mean = float(np.mean(vals))
                    ndti_median = float(np.median(vals))
                    p10, p25, p75, p90 = [float(x) for x in np.percentile(vals, [10, 25, 75, 90])]

                # -------- Guardar máscara (1 banda uint8) --------
                prof_mask = src.profile.copy()
                prof_mask.update(
                    crs=dst_crs,
                    transform=dst_transform,
                    width=dst_width,
                    height=dst_height,
                    count=1,
                    dtype=rasterio.uint8,
                    nodata=0
                )
                with rasterio.open(out_mask_path, "w", **prof_mask) as dst:
                    dst.write((water_mask_dst > 0).astype(np.uint8), 1)

                # -------- Guardar NDTI (1 banda float32) --------
                prof_ndti = src.profile.copy()
                prof_ndti.update(
                    crs=dst_crs,
                    transform=dst_transform,
                    width=dst_width,
                    height=dst_height,
                    count=1,
                    dtype=rasterio.float32
                )
                with rasterio.open(out_ndti_path, "w", **prof_ndti) as dst:
                    dst.write(ndti_water.astype(np.float32), 1)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo guardar/calcular:\n{e}")
            return

        # -------- Registrar en MEDICIONES --------
        nueva = pd.DataFrame([{
            "codigo": codigo,
            "fecha": fecha_archivo,
            "area": area_agua,
            "ndti_media": ndti_mean,
            "ndti_median": ndti_median,
            "p10": p10,
            "p25": p25,
            "p75": p75,
            "p90": p90,
            "ruta_mask": out_mask_path,
            "ruta_ndti": out_ndti_path,
            "ruta_rgb" : output_path
        }])
        df_med = pd.concat([df_med, nueva], ignore_index=True)

        # -------- Guardar Excel --------
        try:
            os.makedirs(os.path.dirname(self.db_path), exist_ok=True)
            # -------- Cargar el archivo Excel existente --------
            wb = openpyxl.load_workbook(self.db_path)

            # Leer la hoja "MEDICIONES"
            if "MEDICIONES" in wb.sheetnames:   
                mediciones_sheet = wb["MEDICIONES"]
            else:
                # Si no existe la hoja "MEDICIONES", crear una nueva
                mediciones_sheet = wb.create_sheet("MEDICIONES")

            # Escribir los nuevos datos en la hoja "MEDICIONES"
            for row in nueva.values.tolist():
                mediciones_sheet.append(row)

            # -------- Guardar el archivo Excel con la hoja "MEDICIONES" actualizada --------
            wb.save(self.db_path)

            QMessageBox.information(
                self, "OK",
                "Registro guardado.\n\n"
                f"Máscara: {out_mask_path}\n"
                f"NDTI: {out_ndti_path}\n"
                f"Área (m²): {area_agua:.2f}"
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo guardar el Excel:\n{e}")
        
        self.CerrarImagen()
    
   
    def CerrarImagen(self):
        
        self.close()
        if self.parent_window:
            self.parent_window.buscar_datos() 
            self.parent_window.show()
    
    def btnAplicar(self):
        
        self.cargar_bandas()
        
        

           
